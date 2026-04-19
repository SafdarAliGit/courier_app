"""
Location API — State/Province and City management
Handles stats, seeding from defaults, template generation, and Excel import/export.
"""

import frappe
from frappe import _
from io import BytesIO


# ─── STATS ──────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_location_stats():
    """Return counts of states/provinces and cities in the database."""
    states   = frappe.db.count("State or Province")
    cities   = frappe.db.count("City")
    ctry_s   = frappe.db.sql(
        "SELECT COUNT(DISTINCT country) FROM `tabState or Province`"
    )[0][0] or 0
    ctry_c   = frappe.db.sql(
        "SELECT COUNT(DISTINCT country) FROM `tabCity`"
    )[0][0] or 0
    return {
        "states":                   states,
        "cities":                   cities,
        "countries_with_states":    int(ctry_s),
        "countries_with_cities":    int(ctry_c),
    }


# ─── SEED FROM DEFAULTS ─────────────────────────────────────────────────────

@frappe.whitelist()
def seed_location_data():
    """
    Populate State or Province and City doctypes from the hardcoded
    _STATES, _STATE_CITIES, _CITIES dictionaries in shipment_api.py.
    Skips records that already exist (safe to run multiple times).
    """
    frappe.only_for("System Manager")

    from courier_app.api.shipment_api import _STATES, _STATE_CITIES, _CITIES

    created_states = 0
    created_cities = 0
    skipped        = 0

    # ── 1. Seed states ──
    for country, states in _STATES.items():
        if not frappe.db.exists("Country", country):
            continue
        for state_name in states:
            doc_name = f"{country}-{state_name}"
            if frappe.db.exists("State or Province", doc_name):
                skipped += 1
                continue
            frappe.get_doc({
                "doctype":    "State or Province",
                "state_name": state_name,
                "country":    country,
                "is_active":  1,
            }).insert(ignore_permissions=True)
            created_states += 1

    frappe.db.commit()

    # ── 2. Seed cities (state-linked from _STATE_CITIES) ──
    for country, state_map in _STATE_CITIES.items():
        if not frappe.db.exists("Country", country):
            continue
        for state_name, cities in state_map.items():
            state_doc = f"{country}-{state_name}"
            state_exists = frappe.db.exists("State or Province", state_doc)
            for city_name in cities:
                doc_name = f"{state_doc}-{city_name}" if state_exists else f"{country}-{city_name}"
                if frappe.db.exists("City", doc_name):
                    skipped += 1
                    continue
                frappe.get_doc({
                    "doctype":           "City",
                    "city_name":         city_name,
                    "country":           country,
                    "state_or_province": state_doc if state_exists else None,
                    "is_active":         1,
                }).insert(ignore_permissions=True)
                created_cities += 1

    frappe.db.commit()

    # ── 3. Seed cities (country-only from _CITIES, skip duplicates) ──
    for country, cities in _CITIES.items():
        if not frappe.db.exists("Country", country):
            continue
        for city_name in cities:
            doc_name = f"{country}-{city_name}"
            if frappe.db.exists("City", doc_name):
                skipped += 1
                continue
            frappe.get_doc({
                "doctype":    "City",
                "city_name":  city_name,
                "country":    country,
                "is_active":  1,
            }).insert(ignore_permissions=True)
            created_cities += 1

    frappe.db.commit()

    return {
        "status":         "success",
        "created_states": created_states,
        "created_cities": created_cities,
        "skipped":        skipped,
    }


# ─── TEMPLATE DOWNLOAD ──────────────────────────────────────────────────────

@frappe.whitelist()
def get_states_template():
    """Stream a blank Excel template for States / Provinces import."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        frappe.throw("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "States"

    headers = ["Country", "State / Province Name", "Active (1=Yes, 0=No)"]
    ws.append(headers)

    # Style header row
    hdr_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    # Sample rows
    for row in [
        ["Pakistan",      "Punjab",      1],
        ["Pakistan",      "Sindh",       1],
        ["United States", "California",  1],
        ["United States", "Texas",       1],
    ]:
        ws.append(row)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22

    buf = BytesIO()
    wb.save(buf)

    frappe.response["filename"]    = "states_provinces_template.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"]        = "binary"


@frappe.whitelist()
def get_cities_template():
    """Stream a blank Excel template for Cities import."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        frappe.throw("openpyxl is required.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cities"

    headers = ["Country", "State / Province (optional)", "City Name", "Active (1=Yes, 0=No)"]
    ws.append(headers)

    hdr_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row in [
        ["Pakistan",      "Punjab",      "Lahore",      1],
        ["Pakistan",      "Punjab",      "Faisalabad",  1],
        ["Pakistan",      "",            "Karachi",     1],
        ["United States", "California",  "Los Angeles", 1],
        ["United States", "",            "Chicago",     1],
    ]:
        ws.append(row)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22

    buf = BytesIO()
    wb.save(buf)

    frappe.response["filename"]    = "cities_template.xlsx"
    frappe.response["filecontent"] = buf.getvalue()
    frappe.response["type"]        = "binary"


# ─── IMPORT ─────────────────────────────────────────────────────────────────

def _resolve_file_path(file_url):
    if "/private/files/" in file_url:
        return frappe.get_site_path("private", "files", file_url.split("/private/files/")[-1])
    return frappe.get_site_path("public", "files", file_url.split("/files/")[-1])


@frappe.whitelist()
def import_states(file_url, mode="upsert"):
    """
    Import States / Provinces from an uploaded Excel file.
    Expected columns: Country | State Name | Active
    """
    frappe.only_for("System Manager")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(_resolve_file_path(file_url), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        frappe.throw(f"Could not read Excel file: {e}")

    created = updated = failed = 0
    errors  = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        country    = str(row[0]).strip() if row[0] else ""
        state_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        try:
            is_active = int(row[2]) if len(row) > 2 and row[2] is not None else 1
        except (ValueError, TypeError):
            is_active = 1

        if not country or not state_name:
            errors.append(f"Row {i}: Country and State Name are required"); failed += 1; continue

        if not frappe.db.exists("Country", country):
            errors.append(f"Row {i}: Country '{country}' not found in Frappe"); failed += 1; continue

        doc_name = f"{country}-{state_name}"
        exists   = frappe.db.exists("State or Province", doc_name)

        if exists and mode == "skip_existing":
            continue

        try:
            if exists and mode == "upsert":
                doc            = frappe.get_doc("State or Province", doc_name)
                doc.is_active  = is_active
                doc.save(ignore_permissions=True)
                updated += 1
            elif not exists:
                frappe.get_doc({
                    "doctype":    "State or Province",
                    "state_name": state_name,
                    "country":    country,
                    "is_active":  is_active,
                }).insert(ignore_permissions=True)
                created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}"); failed += 1

    frappe.db.commit()

    total = created + updated
    return {
        "status":  "success" if not failed else ("partial" if total > 0 else "failed"),
        "created": created,
        "updated": updated,
        "failed":  failed,
        "errors":  errors[:20],
    }


@frappe.whitelist()
def import_cities(file_url, mode="upsert"):
    """
    Import Cities from an uploaded Excel file.
    Expected columns: Country | State/Province (optional) | City Name | Active
    """
    frappe.only_for("System Manager")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(_resolve_file_path(file_url), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        frappe.throw(f"Could not read Excel file: {e}")

    created = updated = failed = 0
    errors  = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        country    = str(row[0]).strip() if row[0] else ""
        state_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        city_name  = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        try:
            is_active = int(row[3]) if len(row) > 3 and row[3] is not None else 1
        except (ValueError, TypeError):
            is_active = 1

        if not country or not city_name:
            errors.append(f"Row {i}: Country and City Name are required"); failed += 1; continue

        if not frappe.db.exists("Country", country):
            errors.append(f"Row {i}: Country '{country}' not found"); failed += 1; continue

        state_doc_name = None
        if state_name:
            state_doc_name = f"{country}-{state_name}"
            if not frappe.db.exists("State or Province", state_doc_name):
                errors.append(
                    f"Row {i}: State '{state_name}' not found for {country}. "
                    "Import states first."
                )
                failed += 1; continue

        doc_name = (
            f"{state_doc_name}-{city_name}" if state_doc_name
            else f"{country}-{city_name}"
        )
        exists = frappe.db.exists("City", doc_name)

        if exists and mode == "skip_existing":
            continue

        try:
            if exists and mode == "upsert":
                doc           = frappe.get_doc("City", doc_name)
                doc.is_active = is_active
                doc.save(ignore_permissions=True)
                updated += 1
            elif not exists:
                frappe.get_doc({
                    "doctype":           "City",
                    "city_name":         city_name,
                    "country":           country,
                    "state_or_province": state_doc_name,
                    "is_active":         is_active,
                }).insert(ignore_permissions=True)
                created += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}"); failed += 1

    frappe.db.commit()

    total = created + updated
    return {
        "status":  "success" if not failed else ("partial" if total > 0 else "failed"),
        "created": created,
        "updated": updated,
        "failed":  failed,
        "errors":  errors[:20],
    }


# ─── INLINE MANAGEMENT (list / add / delete) ────────────────────────────────

@frappe.whitelist()
def list_states(country):
    """Return all states for a country, ordered by name."""
    rows = frappe.get_all(
        "State or Province",
        filters={"country": country},
        fields=["name", "state_name", "is_active"],
        order_by="state_name asc",
        limit=500,
    )
    return rows


@frappe.whitelist()
def add_state(country, state_name):
    """Create a new State or Province."""
    frappe.only_for("System Manager")
    state_name = state_name.strip()
    if not state_name:
        frappe.throw("State name cannot be empty.")
    doc_name = f"{country}-{state_name}"
    if frappe.db.exists("State or Province", doc_name):
        frappe.throw(f"'{state_name}' already exists for {country}.")
    doc = frappe.get_doc({
        "doctype":    "State or Province",
        "state_name": state_name,
        "country":    country,
        "is_active":  1,
    }).insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name, "state_name": doc.state_name, "is_active": 1}


@frappe.whitelist()
def delete_state(name):
    """Delete a State or Province and its linked cities."""
    frappe.only_for("System Manager")
    if not frappe.db.exists("State or Province", name):
        frappe.throw("State not found.")
    cities = frappe.get_all("City", filters={"state_or_province": name}, pluck="name")
    for c in cities:
        frappe.delete_doc("City", c, ignore_permissions=True)
    frappe.delete_doc("State or Province", name, ignore_permissions=True)
    frappe.db.commit()
    return {"deleted": name, "cities_removed": len(cities)}


@frappe.whitelist()
def list_cities(country, state_or_province=""):
    """Return cities for a country, optionally filtered by state."""
    filters = {"country": country}
    if state_or_province:
        filters["state_or_province"] = state_or_province
    rows = frappe.get_all(
        "City",
        filters=filters,
        fields=["name", "city_name", "state_or_province", "is_active"],
        order_by="city_name asc",
        limit=1000,
    )
    return rows


@frappe.whitelist()
def add_city(country, city_name, state_or_province=""):
    """Create a new City record."""
    frappe.only_for("System Manager")
    city_name = city_name.strip()
    if not city_name:
        frappe.throw("City name cannot be empty.")
    state_or_province = (state_or_province or "").strip() or None
    doc_name = f"{state_or_province}-{city_name}" if state_or_province else f"{country}-{city_name}"
    if frappe.db.exists("City", doc_name):
        frappe.throw(f"'{city_name}' already exists.")
    doc = frappe.get_doc({
        "doctype":           "City",
        "city_name":         city_name,
        "country":           country,
        "state_or_province": state_or_province,
        "is_active":         1,
    }).insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name, "city_name": doc.city_name, "state_or_province": state_or_province, "is_active": 1}


@frappe.whitelist()
def delete_city(name):
    """Delete a City record."""
    frappe.only_for("System Manager")
    if not frappe.db.exists("City", name):
        frappe.throw("City not found.")
    frappe.delete_doc("City", name, ignore_permissions=True)
    frappe.db.commit()
    return {"deleted": name}
